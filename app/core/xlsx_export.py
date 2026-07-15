from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Issue
from app.schemas import ProjectStatsOut, SprintStatsOut, UserReportOut


def _autofit(ws: Worksheet) -> None:
    for col in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 60)


def _header_row(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_project_stats_workbook(stats: ProjectStatsOut) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _header_row(ws, ["Metric", "Value"])
    ws.append(["Project", stats.project_key])
    ws.append(["Range", stats.range])
    ws.append(["Range Start", stats.range_start.isoformat()])
    ws.append(["Range End", stats.range_end.isoformat()])
    ws.append(["Total Issues", stats.total_issues])
    ws.append(["Done", stats.done_count])
    ws.append(["Open", stats.open_count])
    ws.append(["Total Sprints", stats.total_sprints])
    ws.append(["Active Sprints", stats.active_sprint_count])
    ws.append(["Total Users", stats.total_users])
    ws.append(["Active Users", stats.active_users])
    ws.append(["Total Hours Logged", stats.total_hours_logged])
    ws.append(["Total Original Estimate (h)", stats.total_original_estimate_hours])
    ws.append(["Total Remaining Estimate (h)", stats.total_remaining_estimate_hours])
    ws.append(["Issues Created In Range", stats.issues_created_in_range])
    ws.append(["Issues Completed In Range", stats.issues_completed_in_range])

    ws2 = wb.create_sheet("Status Breakdown")
    _header_row(ws2, ["Status", "Count"])
    for row in stats.status_breakdown:
        ws2.append([row.status, row.count])

    ws3 = wb.create_sheet("Issue Type Breakdown")
    _header_row(ws3, ["Issue Type", "Count"])
    for row in stats.type_breakdown:
        ws3.append([row.issue_type, row.count])

    ws4 = wb.create_sheet("Hours Logged Over Time")
    _header_row(ws4, ["Date", "Hours"])
    for point in stats.hours_logged_series:
        ws4.append([point.date.isoformat(), round(point.minutes / 60, 2)])

    for sheet in wb.worksheets:
        _autofit(sheet)
    return workbook_to_bytes(wb)


def build_sprint_issues_workbook(sprint_name: str, issues: list[Issue]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"[:31]
    headers = [
        "Issue Key",
        "Title",
        "Description",
        "Type",
        "Priority",
        "Status",
        "Assignee",
        "Reporter",
        "Story Points",
        "Original Estimate (h)",
        "Remaining Estimate (h)",
        "Time Logged (h)",
        "Due Date",
    ]
    _header_row(ws, headers)
    for issue in issues:
        time_logged_minutes = sum(w.time_spent_minutes for w in issue.worklogs)
        ws.append(
            [
                issue.issue_key,
                issue.title,
                issue.description or "",
                issue.issue_type.value,
                issue.priority.value,
                issue.status,
                issue.assignee.name if issue.assignee else "",
                issue.reporter.name if issue.reporter else "",
                issue.story_points if issue.story_points is not None else "",
                round(issue.original_estimate_minutes / 60, 2)
                if issue.original_estimate_minutes is not None
                else "",
                round(issue.remaining_estimate_minutes / 60, 2)
                if issue.remaining_estimate_minutes is not None
                else "",
                round(time_logged_minutes / 60, 2),
                issue.due_date.isoformat() if issue.due_date else "",
            ]
        )
    _autofit(ws)
    return workbook_to_bytes(wb)


def build_user_report_workbook(report: UserReportOut) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "User Report"
    headers = [
        "Name",
        "Email",
        "Hours Logged",
        "Tickets Assigned",
        "Tickets Completed",
        "Estimate Hours",
        "Assigned Ticket Keys",
        "Completed Ticket Keys",
    ]
    _header_row(ws, headers)
    for row in report.users:
        ws.append(
            [
                row.name,
                row.email,
                row.hours_logged,
                row.tickets_assigned_count,
                row.tickets_completed_count,
                row.estimate_hours,
                ", ".join(row.assigned_ticket_keys),
                ", ".join(row.completed_ticket_keys),
            ]
        )
    _autofit(ws)
    return workbook_to_bytes(wb)
